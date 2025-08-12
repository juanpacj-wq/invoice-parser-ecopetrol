"""
Módulo para la exportación de datos de facturas individuales a formato Excel.
Este módulo contiene la clase ExportadorExcel para generar archivos Excel
con múltiples hojas a partir de los datos de una factura.
"""

import uuid
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ExportadorExcel:
    """
    Clase para exportar datos de facturas a formato Excel.
    Esta clase genera archivos Excel con múltiples hojas, estilos y metadatos.
    """
    
    def __init__(self, datos_procesados, ruta_salida=None):
        """
        Inicializa el exportador con los datos procesados.
        
        Args:
            datos_procesados (dict): Datos procesados de la factura
            ruta_salida (str, optional): Ruta donde se guardará el archivo Excel.
                                         Si es None, se usará 'factura_analizada.xlsx'.
        """
        self.datos_procesados = datos_procesados
        self.ruta_salida = ruta_salida or "factura_analizada.xlsx"
        self.invoice_id = str(uuid.uuid4())
        self.workbook = Workbook()
        
        # Definir estilos
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
    
    def exportar(self):
        """
        Exporta los datos procesados a un archivo Excel.
        Crea las hojas de 'Facturas', 'Conceptos', 'Autorizaciones', 'Parámetros Específicos' y 'Metadatos'.
        
        Returns:
            str: Ruta del archivo Excel creado
        """
        # Crear hoja de Facturas
        self._crear_hoja_facturas()
        
        # Crear hoja de Conceptos
        self._crear_hoja_conceptos()
        
        # Crear hoja de Autorizaciones
        self._crear_hoja_autorizaciones()
        
        # Crear hoja de Parámetros Específicos
        self._crear_hoja_parametros_especificos()
        
        # Crear hoja de Metadatos
        self._crear_hoja_metadatos()
        
        # Guardar archivo
        self.workbook.save(self.ruta_salida)
        return self.ruta_salida
    
    def _crear_hoja_facturas(self):
        """
        Crea la hoja 'Facturas' con los datos generales de la factura.
        """
        ws_facturas = self.workbook.active
        ws_facturas.title = "Facturas"
        
        # Encabezados de la tabla de Facturas
        encabezados_factura = [
            "ID_Factura", "Numero_Factura", "Fecha_Procesamiento", "Fecha_Vencimiento", "Periodo_Facturacion", 
            "Factor_M", "Codigo_SIC", "Subtotal_Base_Energia", "Contribucion", 
            "Contribucion_Otros_Meses", "Subtotal_Energia_Contribucion_KWh", 
            "Subtotal_Energia_Contribucion_Pesos", "Otros_Cobros", "Sobretasa", 
            "Ajustes_Cargos_Regulados", "Compensaciones", "Saldo_Cartera", 
            "Interes_Mora","Recobros", "Alumbrado_Publico", "Impuesto_Alumbrado_Publico", 
            "Ajuste_IAP_Otros_Meses", "Convivencia_Ciudadana", "Tasa_Especial_Convivencia", 
            "Ajuste_Tasa_Convivencia", "Total_Servicio_Energia_Impuestos", 
            "Ajuste_Decena", "Neto_Pagar", "Energia_Reactiva_Inductiva", 
            "Energia_Reactiva_Capacitiva", "Total_Energia_Reactiva"
        ]
        
        # Aplicar encabezados y estilos
        for col_idx, header in enumerate(encabezados_factura, 1):
            cell = ws_facturas.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Obtener datos generales
        datos_generales = self.datos_procesados["datos_generales"]
        
        # Añadir datos de factura
        factura_valores = [
            self.invoice_id,
            datos_generales.get("numero_factura", "No encontrado"),
            self.datos_procesados.get("fecha_procesamiento", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            datos_generales.get("fecha_vencimiento", "No encontrado"),
            datos_generales.get("periodo_facturacion", "No encontrado"),
            datos_generales.get("factor_m", "No encontrado"),
            datos_generales.get("codigo_sic", "No encontrado"),
            datos_generales.get("subtotal_base_energia", 0),
            datos_generales.get("contribucion", 0),
            datos_generales.get("contribucion_otros_meses", 0),
            datos_generales.get("subtotal_energia_contribucion_kwh", 0),
            datos_generales.get("subtotal_energia_contribucion_pesos", 0),
            datos_generales.get("otros_cobros", 0),
            datos_generales.get("sobretasa", 0),
            datos_generales.get("ajustes_cargos_regulados", 0),
            datos_generales.get("compensaciones", 0),
            datos_generales.get("saldo_cartera", 0),
            datos_generales.get("interes_mora", 0),
            datos_generales.get("recobros", 0),
            datos_generales.get("alumbrado_publico", 0),
            datos_generales.get("impuesto_alumbrado_publico", 0),
            datos_generales.get("ajuste_iap_otros_meses", 0),
            datos_generales.get("convivencia_ciudadana", 0),
            datos_generales.get("tasa_especial_convivencia", 0),
            datos_generales.get("ajuste_tasa_convivencia", 0),
            datos_generales.get("total_servicio_energia_impuestos", 0),
            datos_generales.get("ajuste_decena", 0),
            datos_generales.get("neto_pagar", 0),
            datos_generales.get("energia_reactiva_inductiva", 0),
            datos_generales.get("energia_reactiva_capacitiva", 0),
            datos_generales.get("total_energia_reactiva", 0)
        ]
        
        for col_idx, value in enumerate(factura_valores, 1):
            cell = ws_facturas.cell(row=2, column=col_idx, value=value)
            cell.border = self.border
        
        # Ajustar ancho de columnas
        self._ajustar_ancho_columnas(ws_facturas)
    
    def _crear_hoja_conceptos(self):
        """
        Crea la hoja 'Conceptos' con los datos de los componentes de la factura.
        """
        ws_conceptos = self.workbook.create_sheet(title="Conceptos")
        
        # Encabezados de la tabla de Conceptos
        headers_conceptos = ["ID_Factura", "Codigo_SIC", "Concepto", "KWh_KVArh", "Precio_KWh", "Mes_Corriente", "Mes_Anteriores", "Total"]
        
        for col_idx, header in enumerate(headers_conceptos, 1):
            cell = ws_conceptos.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Añadir datos de componentes
        componentes = self.datos_procesados["componentes"]
        
        # Obtener el código SIC de los datos generales
        codigo_sic = self.datos_procesados["datos_generales"].get("codigo_sic", "No encontrado")
        
        for row_idx, componente in enumerate(componentes, 2):
            # Añadir ID de factura a cada componente
            ws_conceptos.cell(row=row_idx, column=1, value=self.invoice_id).border = self.border
            
            # Añadir Código SIC a cada componente
            ws_conceptos.cell(row=row_idx, column=2, value=codigo_sic).border = self.border
            
            # Datos del componente
            ws_conceptos.cell(row=row_idx, column=3, value=componente.get("concepto", "")).border = self.border
            ws_conceptos.cell(row=row_idx, column=4, value=componente.get("kwh_kvarh", "N/A")).border = self.border
            ws_conceptos.cell(row=row_idx, column=5, value=componente.get("precio_kwh", 0)).border = self.border
            ws_conceptos.cell(row=row_idx, column=6, value=componente.get("mes_corriente", 0)).border = self.border
            ws_conceptos.cell(row=row_idx, column=7, value=componente.get("mes_anteriores", 0)).border = self.border
            ws_conceptos.cell(row=row_idx, column=8, value=componente.get("total", 0)).border = self.border
        
        # Ajustar ancho de columnas
        self._ajustar_ancho_columnas(ws_conceptos)
    
    def _crear_hoja_autorizaciones(self):
        """
        Crea la hoja 'Autorizaciones' con los valores HES de la factura.
        """
        ws_autorizaciones = self.workbook.create_sheet(title="Autorizaciones")
        
        # Encabezados de la tabla de Autorizaciones
        headers_autorizaciones = [
            "ID_Factura", "Codigo_SIC", "HES1", "HES2", "HES3", "HES4", "HES5", 
            "HES6", "HES7", "HES8", "HES9", "HES10"
        ]
        
        for col_idx, header in enumerate(headers_autorizaciones, 1):
            cell = ws_autorizaciones.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Obtener datos generales
        datos_generales = self.datos_procesados["datos_generales"]
        codigo_sic = datos_generales.get("codigo_sic", "No encontrado")
        
        # Añadir fila con valores HES
        hes_values = [
            self.invoice_id,
            codigo_sic,
            datos_generales.get("hes1", ""),
            datos_generales.get("hes2", ""),
            datos_generales.get("hes3", ""),
            datos_generales.get("hes4", ""),
            datos_generales.get("hes5", ""),
            datos_generales.get("hes6", ""),
            datos_generales.get("hes7", ""),
            datos_generales.get("hes8", ""),
            datos_generales.get("hes9", ""),
            datos_generales.get("hes10", "")
        ]
        
        for col_idx, value in enumerate(hes_values, 1):
            cell = ws_autorizaciones.cell(row=2, column=col_idx, value=value)
            cell.border = self.border
        
        # Ajustar ancho de columnas
        self._ajustar_ancho_columnas(ws_autorizaciones)
    
    def _crear_hoja_parametros_especificos(self):
        """
        Crea la hoja 'Parametros' con los valores específicos extraídos de la factura.
        """
        ws_parametros = self.workbook.create_sheet(title="Datos_OR")
        
        # Encabezados de la tabla de Parámetros Específicos
        headers_parametros = [
            "ID_Factura", "Codigo_SIC", "IR", "Grupo", "DIU_INT", "DIUM_INT", 
            "FIU_INT", "FIUM_INT", "FIUG", "DIUG"
        ]
        
        for col_idx, header in enumerate(headers_parametros, 1):
            cell = ws_parametros.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Obtener datos generales y parámetros específicos
        datos_generales = self.datos_procesados["datos_generales"]
        parametros_especificos = self.datos_procesados["parametros_especificos"]
        codigo_sic = datos_generales.get("codigo_sic", "No encontrado")
        
        # Añadir fila con valores de parámetros específicos
        parametros_valores = [
            self.invoice_id,
            codigo_sic,
            parametros_especificos.get("ir", 0),
            parametros_especificos.get("grupo", 0),
            parametros_especificos.get("diu_int", 0),
            parametros_especificos.get("dium_int", 0),
            parametros_especificos.get("fiu_int", 0),
            parametros_especificos.get("fium_int", 0),
            parametros_especificos.get("fiug", 0),
            parametros_especificos.get("diug", 0)
        ]
        
        for col_idx, value in enumerate(parametros_valores, 1):
            cell = ws_parametros.cell(row=2, column=col_idx, value=value)
            cell.border = self.border
        
        # Ajustar ancho de columnas
        self._ajustar_ancho_columnas(ws_parametros)
    
    def _crear_hoja_metadatos(self):
        """
        Crea la hoja 'Metadatos' con información descriptiva de los campos.
        """
        ws_metadata = self.workbook.create_sheet(title="Metadatos")
        
        metadata_headers = ["Campo", "Descripción", "Tipo de Dato", "Unidad", "Observaciones"]
        
        for col_idx, header in enumerate(metadata_headers, 1):
            cell = ws_metadata.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Añadir descripciones de metadatos para tabla Facturas
        filas_metadatos = [
            ["ID_Factura", "Identificador único de la factura", "UUID", "N/A", "Clave primaria"],
            ["Numero_Factura", "Número de la factura electrónica", "String", "N/A", "Ej. E1888"],
            ["Fecha_Procesamiento", "Fecha y hora de procesamiento del archivo", "Datetime", "N/A", "Generado automáticamente"],
            ["Fecha_Vencimiento", "Fecha límite de pago", "Date", "N/A", ""],
            ["Periodo_Facturacion", "Período al que corresponde la factura", "DATE", "N/A", ""],
            ["Factor_M", "Factor de multiplicación para mediciones", "Integer", "N/A", ""],
            ["Codigo_SIC", "Código del Sistema de Información Comercial", "String", "N/A", ""],
            ["Subtotal_Base_Energia", "Subtotal base de energía facturada", "Decimal", "Pesos", ""],
            ["Contribucion", "Valor de la contribución", "Decimal", "Pesos", ""],
            ["Contribucion_Otros_Meses", "Contribución correspondiente a meses anteriores", "Decimal", "Pesos", ""],
            ["Subtotal_Energia_Contribucion_KWh", "Precio por kWh de energía y contribución", "Decimal", "$/kWh", ""],
            ["Subtotal_Energia_Contribucion_Pesos", "Subtotal energía más contribución", "Decimal", "Pesos", ""],
            ["Otros_Cobros", "Valor de otros cobros adicionales", "Decimal", "Pesos", ""],
            ["Sobretasa", "Valor de la sobretasa aplicada", "Decimal", "Pesos", ""],
            ["Ajustes_Cargos_Regulados", "Ajustes por cargos regulados", "Decimal", "Pesos", ""],
            ["Compensaciones", "Valor de compensaciones aplicadas", "Decimal", "Pesos", ""],
            ["Saldo_Cartera", "Saldo pendiente en cartera", "Decimal", "Pesos", ""],
            ["Interes_Mora", "Interés por mora en pagos anteriores", "Decimal", "Pesos", ""],
            ["Recobros", "Trabajos otc", "Decimal", "Pesos", ""],
            ["Alumbrado_Publico", "Cargo por alumbrado público", "Decimal", "Pesos", ""],
            ["Impuesto_Alumbrado_Publico", "Impuesto por alumbrado público", "Decimal", "Pesos", ""],
            ["Ajuste_IAP_Otros_Meses", "Ajuste de impuesto de alumbrado público de otros meses", "Decimal", "Pesos", ""],
            ["Convivencia_Ciudadana", "Cargo por convivencia ciudadana", "Decimal", "Pesos", ""],
            ["Tasa_Especial_Convivencia", "Tasa especial de convivencia ciudadana", "Decimal", "Pesos", ""],
            ["Ajuste_Tasa_Convivencia", "Ajuste de la tasa de convivencia de otros meses", "Decimal", "Pesos", ""],
            ["Total_Servicio_Energia_Impuestos", "Total por servicio de energía e impuestos", "Decimal", "Pesos", ""],
            ["Ajuste_Decena", "Ajuste a la decena", "Decimal", "Pesos", ""],
            ["Neto_Pagar", "Valor neto a pagar", "Decimal", "Pesos", ""],
            ["Energia_Reactiva_Inductiva", "Valor de energía reactiva inductiva", "Decimal", "kVArh", ""],
            ["Energia_Reactiva_Capacitiva", "Valor de energía reactiva capacitiva", "Decimal", "kVArh", ""],
            ["Total_Energia_Reactiva", "Total de energía reactiva", "Decimal", "kVArh", ""]
        ]
        
        # Añadir metadatos para tabla Conceptos
        conceptos_metadatos = [
            ["ID_Factura", "Identificador único de la factura (clave foránea)", "UUID", "N/A", "Relaciona con tabla Facturas"],
            ["Codigo_SIC", "Código del Sistema de Información Comercial", "String", "N/A", "Identifica la cuenta en el sistema"],
            ["Concepto", "Tipo de concepto facturado", "String", "N/A", "Ej. Generación, Comercialización"],
            ["KWh_KVArh", "Cantidad de energía consumida", "Decimal", "kWh/kVArh", "Solo aplica para algunos conceptos"],
            ["Precio_KWh", "Precio unitario por kWh", "Decimal", "$/kWh", ""],
            ["Mes_Corriente", "Valor facturado del mes actual", "Decimal", "Pesos", ""],
            ["Mes_Anteriores", "Valor facturado de meses anteriores", "Decimal", "Pesos", ""],
            ["Total", "Valor total del concepto", "Decimal", "Pesos", ""]
        ]
        
        # Añadir metadatos para tabla Autorizaciones
        autorizaciones_metadatos = [
            ["ID_Factura", "Identificador único de la factura (clave foránea)", "UUID", "N/A", "Relaciona con tabla Facturas"],
            ["Codigo_SIC", "Código del Sistema de Información Comercial", "String", "N/A", "Identifica la cuenta en el sistema"],
            ["HES1", "Autorización 1", "Integer", "N/A", "Código de autorización 1"],
            ["HES2", "Autorización 2", "Integer", "N/A", "Código de autorización 2"],
            ["HES3", "Autorización 3", "Integer", "N/A", "Código de autorización 3"],
            ["HES4", "Autorización 4", "Integer", "N/A", "Código de autorización 4"],
            ["HES5", "Autorización 5", "Integer", "N/A", "Código de autorización 5"],
            ["HES6", "Autorización 6", "Integer", "N/A", "Código de autorización 6"],
            ["HES7", "Autorización 7", "Integer", "N/A", "Código de autorización 7"],
            ["HES8", "Autorización 8", "Integer", "N/A", "Código de autorización 8"],
            ["HES9", "Autorización 9", "Integer", "N/A", "Código de autorización 9"],
            ["HES10", "Autorización 10", "Integer", "N/A", "Código de autorización 10"]
        ]
        
        # Añadir metadatos para tabla Parámetros Específicos
        parametros_metadatos = [
            ["ID_Factura", "Identificador único de la factura (clave foránea)", "UUID", "N/A", "Relaciona con tabla Facturas"],
            ["Codigo_SIC", "Código del Sistema de Información Comercial", "String", "N/A", "Identifica la cuenta en el sistema"],
            ["IR", "Parámetro IR", "Decimal", "N/A", "Valor numérico del parámetro IR"],
            ["Grupo", "Número de grupo", "Decimal", "N/A", "Identificador del grupo"],
            ["DIU_INT", "Parámetro DIU INT", "Decimal", "N/A", "Valor numérico del parámetro DIU INT"],
            ["DIUM_INT", "Parámetro DIUM INT", "Decimal", "N/A", "Valor numérico del parámetro DIUM INT"],
            ["FIU_INT", "Parámetro FIU INT", "Decimal", "N/A", "Valor numérico del parámetro FIU INT"],
            ["FIUM_INT", "Parámetro FIUM INT", "Decimal", "N/A", "Valor numérico del parámetro FIUM INT"],
            ["FIUG", "Parámetro FIUG", "Decimal", "N/A", "Valor decimal del parámetro FIUG"],
            ["DIUG", "Parámetro DIUG", "Decimal", "N/A", "Valor decimal del parámetro DIUG"]
        ]
        
        # Combinar todos los metadatos
        all_metadata = filas_metadatos + conceptos_metadatos + autorizaciones_metadatos + parametros_metadatos
        
        for row_idx, metadata in enumerate(all_metadata, 2):
            for col_idx, value in enumerate(metadata, 1):
                cell = ws_metadata.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border
        
        # Ajustar ancho de columnas
        self._ajustar_ancho_columnas(ws_metadata)
    
    def _ajustar_ancho_columnas(self, worksheet):
        """
        Ajusta el ancho de las columnas de una hoja de cálculo.
        
        Args:
            worksheet: Hoja de cálculo a ajustar
        """
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width
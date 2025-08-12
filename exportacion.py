"""
Módulo para la exportación de datos procesados de facturas a formatos específicos.
Este módulo contiene clases y funciones para exportar los datos procesados a 
formatos como Excel, incluyendo estilos y metadatos.
"""

import uuid
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ExportadorExcel:
    """
    Clase para exportar datos de facturas a formato Excel.
    Esta clase genera archivos Excel con múltiples hojas, estilos y metadatos.
    """
    
    def __init__(self, datos_procesados, ruta_salida=None):
        """
        Inicializa el exportador con los datos procesados.
        
        Args:
            datos_procesados (dict): Datos procesados de la factura
            ruta_salida (str, optional): Ruta donde se guardará el archivo Excel.
                                         Si es None, se usará 'factura_analizada.xlsx'.
        """
        self.datos_procesados = datos_procesados
        self.ruta_salida = ruta_salida or "factura_analizada.xlsx"
        self.invoice_id = str(uuid.uuid4())
        self.workbook = Workbook()
        
        # Definir estilos
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
    
    def exportar(self):
        """
        Exporta los datos procesados a un archivo Excel.
        Crea las hojas de 'Facturas', 'Conceptos', 'Autorizaciones', 'Parámetros Específicos' y 'Metadatos'.
        
        Returns:
            str: Ruta del archivo Excel creado
        """
        # Crear hoja de Facturas
        self._crear_hoja_facturas()
        
        # Crear hoja de Conceptos
        self._crear_hoja_conceptos()
        
        # Crear hoja de Autorizaciones
        self._crear_hoja_autorizaciones()
        
        # Crear hoja de Parámetros Específicos
        self._crear_hoja_parametros_especificos()
        
        # Crear hoja de Metadatos
        self._crear_hoja_metadatos()
        
        # Guardar archivo
        self.workbook.save(self.ruta_salida)
        return self.ruta_salida
    
    def _crear_hoja_facturas(self):
        """
        Crea la hoja 'Facturas' con los datos generales de la factura.
        """
        ws_facturas = self.workbook.active
        ws_facturas.title = "Facturas"
        
        # Encabezados de la tabla de Facturas
        encabezados_factura = [
            "ID_Factura", "Numero_Factura", "Fecha_Procesamiento", "Fecha_Vencimiento", "Periodo_Facturacion", 
            "Factor_M", "Codigo_SIC", "Subtotal_Base_Energia", "Contribucion", 
            "Contribucion_Otros_Meses", "Subtotal_Energia_Contribucion_KWh", 
            "Subtotal_Energia_Contribucion_Pesos", "Otros_Cobros", "Sobretasa", 
            "Ajustes_Cargos_Regulados", "Compensaciones", "Saldo_Cartera", 
            "Interes_Mora","Recobros", "Alumbrado_Publico", "Impuesto_Alumbrado_Publico", 
            "Ajuste_IAP_Otros_Meses", "Convivencia_Ciudadana", "Tasa_Especial_Convivencia", 
            "Ajuste_Tasa_Convivencia", "Total_Servicio_Energia_Impuestos", 
            "Ajuste_Decena", "Neto_Pagar", "Energia_Reactiva_Inductiva", 
            "Energia_Reactiva_Capacitiva", "Total_Energia_Reactiva"
        ]
        
        # Aplicar encabezados y estilos
        for col_idx, header in enumerate(encabezados_factura, 1):
            cell = ws_facturas.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Obtener datos generales
        datos_generales = self.datos_procesados["datos_generales"]
        
        # Añadir datos de factura
        factura_valores = [
            self.invoice_id,
            datos_generales.get("numero_factura", "No encontrado"),
            self.datos_procesados.get("fecha_procesamiento", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            datos_generales.get("fecha_vencimiento", "No encontrado"),
            datos_generales.get("periodo_facturacion", "No encontrado"),
            datos_generales.get("factor_m", "No encontrado"),
            datos_generales.get("codigo_sic", "No encontrado"),
            datos_generales.get("subtotal_base_energia", 0),
            datos_generales.get("contribucion", 0),
            datos_generales.get("contribucion_otros_meses", 0),
            datos_generales.get("subtotal_energia_contribucion_kwh", 0),
            datos_generales.get("subtotal_energia_contribucion_pesos", 0),
            datos_generales.get("otros_cobros", 0),
            datos_generales.get("sobretasa", 0),
            datos_generales.get("ajustes_cargos_regulados", 0),
            datos_generales.get("compensaciones", 0),
            datos_generales.get("saldo_cartera", 0),
            datos_generales.get("interes_mora", 0),
            datos_generales.get("recobros", 0),
            datos_generales.get("alumbrado_publico", 0),
            datos_generales.get("impuesto_alumbrado_publico", 0),
            datos_generales.get("ajuste_iap_otros_meses", 0),
            datos_generales.get("convivencia_ciudadana", 0),
            datos_generales.get("tasa_especial_convivencia", 0),
            datos_generales.get("ajuste_tasa_convivencia", 0),
            datos_generales.get("total_servicio_energia_impuestos", 0),
            datos_generales.get("ajuste_decena", 0),
            datos_generales.get("neto_pagar", 0),
            datos_generales.get("energia_reactiva_inductiva", 0),
            datos_generales.get("energia_reactiva_capacitiva", 0),
            datos_generales.get("total_energia_reactiva", 0)
        ]
        
        for col_idx, value in enumerate(factura_valores, 1):
            cell = ws_facturas.cell(row=2, column=col_idx, value=value)
            cell.border = self.border
        
        # Ajustar ancho de columnas
        self._ajustar_ancho_columnas(ws_facturas)
    
    def _crear_hoja_conceptos(self):
        """
        Crea la hoja 'Conceptos' con los datos de los componentes de la factura.
        """
        ws_conceptos = self.workbook.create_sheet(title="Conceptos")
        
        # Encabezados de la tabla de Conceptos
        headers_conceptos = ["ID_Factura", "Codigo_SIC", "Concepto", "KWh_KVArh", "Precio_KWh", "Mes_Corriente", "Mes_Anteriores", "Total"]
        
        for col_idx, header in enumerate(headers_conceptos, 1):
            cell = ws_conceptos.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Añadir datos de componentes
        componentes = self.datos_procesados["componentes"]
        
        # Obtener el código SIC de los datos generales
        codigo_sic = self.datos_procesados["datos_generales"].get("codigo_sic", "No encontrado")
        
        for row_idx, componente in enumerate(componentes, 2):
            # Añadir ID de factura a cada componente
            ws_conceptos.cell(row=row_idx, column=1, value=self.invoice_id).border = self.border
            
            # Añadir Código SIC a cada componente
            ws_conceptos.cell(row=row_idx, column=2, value=codigo_sic).border = self.border
            
            # Datos del componente
            ws_conceptos.cell(row=row_idx, column=3, value=componente.get("concepto", "")).border = self.border
            ws_conceptos.cell(row=row_idx, column=4, value=componente.get("kwh_kvarh", "N/A")).border = self.border
            ws_conceptos.cell(row=row_idx, column=5, value=componente.get("precio_kwh", 0)).border = self.border
            ws_conceptos.cell(row=row_idx, column=6, value=componente.get("mes_corriente", 0)).border = self.border
            ws_conceptos.cell(row=row_idx, column=7, value=componente.get("mes_anteriores", 0)).border = self.border
            ws_conceptos.cell(row=row_idx, column=8, value=componente.get("total", 0)).border = self.border
        
        # Ajustar ancho de columnas
        self._ajustar_ancho_columnas(ws_conceptos)
    
    def _crear_hoja_autorizaciones(self):
        """
        Crea la hoja 'Autorizaciones' con los valores HES de la factura.
        """
        ws_autorizaciones = self.workbook.create_sheet(title="Autorizaciones")
        
        # Encabezados de la tabla de Autorizaciones
        headers_autorizaciones = [
            "ID_Factura", "Codigo_SIC", "HES1", "HES2", "HES3", "HES4", "HES5", 
            "HES6", "HES7", "HES8", "HES9", "HES10"
        ]
        
        for col_idx, header in enumerate(headers_autorizaciones, 1):
            cell = ws_autorizaciones.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Obtener datos generales
        datos_generales = self.datos_procesados["datos_generales"]
        codigo_sic = datos_generales.get("codigo_sic", "No encontrado")
        
        # Añadir fila con valores HES
        hes_values = [
            self.invoice_id,
            codigo_sic,
            datos_generales.get("hes1", ""),
            datos_generales.get("hes2", ""),
            datos_generales.get("hes3", ""),
            datos_generales.get("hes4", ""),
            datos_generales.get("hes5", ""),
            datos_generales.get("hes6", ""),
            datos_generales.get("hes7", ""),
            datos_generales.get("hes8", ""),
            datos_generales.get("hes9", ""),
            datos_generales.get("hes10", "")
        ]
        
        for col_idx, value in enumerate(hes_values, 1):
            cell = ws_autorizaciones.cell(row=2, column=col_idx, value=value)
            cell.border = self.border
        
        # Ajustar ancho de columnas
        self._ajustar_ancho_columnas(ws_autorizaciones)
    
    def _crear_hoja_parametros_especificos(self):
        """
        Crea la hoja 'Parametros' con los valores específicos extraídos de la factura.
        """
        ws_parametros = self.workbook.create_sheet(title="Datos_OR")
        
        # Encabezados de la tabla de Parámetros Específicos
        headers_parametros = [
            "ID_Factura", "Codigo_SIC", "IR", "Grupo", "DIU_INT", "DIUM_INT", 
            "FIU_INT", "FIUM_INT", "FIUG", "DIUG"
        ]
        
        for col_idx, header in enumerate(headers_parametros, 1):
            cell = ws_parametros.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Obtener datos generales y parámetros específicos
        datos_generales = self.datos_procesados["datos_generales"]
        parametros_especificos = self.datos_procesados["parametros_especificos"]
        codigo_sic = datos_generales.get("codigo_sic", "No encontrado")
        
        # Añadir fila con valores de parámetros específicos
        parametros_valores = [
            self.invoice_id,
            codigo_sic,
            parametros_especificos.get("ir", 0),
            parametros_especificos.get("grupo", 0),
            parametros_especificos.get("diu_int", 0),
            parametros_especificos.get("dium_int", 0),
            parametros_especificos.get("fiu_int", 0),
            parametros_especificos.get("fium_int", 0),
            parametros_especificos.get("fiug", 0),
            parametros_especificos.get("diug", 0)
        ]
        
        for col_idx, value in enumerate(parametros_valores, 1):
            cell = ws_parametros.cell(row=2, column=col_idx, value=value)
            cell.border = self.border
        
        # Ajustar ancho de columnas
        self._ajustar_ancho_columnas(ws_parametros)
    
    def _crear_hoja_metadatos(self):
        """
        Crea la hoja 'Metadatos' con información descriptiva de los campos.
        """
        ws_metadata = self.workbook.create_sheet(title="Metadatos")
        
        metadata_headers = ["Campo", "Descripción", "Tipo de Dato", "Unidad", "Observaciones"]
        
        for col_idx, header in enumerate(metadata_headers, 1):
            cell = ws_metadata.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Añadir descripciones de metadatos para tabla Facturas
        filas_metadatos = [
            ["ID_Factura", "Identificador único de la factura", "UUID", "N/A", "Clave primaria"],
            ["Numero_Factura", "Número de la factura electrónica", "String", "N/A", "Ej. E1888"],
            ["Fecha_Procesamiento", "Fecha y hora de procesamiento del archivo", "Datetime", "N/A", "Generado automáticamente"],
            ["Fecha_Vencimiento", "Fecha límite de pago", "Date", "N/A", ""],
            ["Periodo_Facturacion", "Período al que corresponde la factura", "DATE", "N/A", ""],
            ["Factor_M", "Factor de multiplicación para mediciones", "Integer", "N/A", ""],
            ["Codigo_SIC", "Código del Sistema de Información Comercial", "String", "N/A", ""],
            ["Subtotal_Base_Energia", "Subtotal base de energía facturada", "Decimal", "Pesos", ""],
            ["Contribucion", "Valor de la contribución", "Decimal", "Pesos", ""],
            ["Contribucion_Otros_Meses", "Contribución correspondiente a meses anteriores", "Decimal", "Pesos", ""],
            ["Subtotal_Energia_Contribucion_KWh", "Precio por kWh de energía y contribución", "Decimal", "$/kWh", ""],
            ["Subtotal_Energia_Contribucion_Pesos", "Subtotal energía más contribución", "Decimal", "Pesos", ""],
            ["Otros_Cobros", "Valor de otros cobros adicionales", "Decimal", "Pesos", ""],
            ["Sobretasa", "Valor de la sobretasa aplicada", "Decimal", "Pesos", ""],
            ["Ajustes_Cargos_Regulados", "Ajustes por cargos regulados", "Decimal", "Pesos", ""],
            ["Compensaciones", "Valor de compensaciones aplicadas", "Decimal", "Pesos", ""],
            ["Saldo_Cartera", "Saldo pendiente en cartera", "Decimal", "Pesos", ""],
            ["Interes_Mora", "Interés por mora en pagos anteriores", "Decimal", "Pesos", ""],
            ["Recobros", "Trabajos otc", "Decimal", "Pesos", ""],
            ["Alumbrado_Publico", "Cargo por alumbrado público", "Decimal", "Pesos", ""],
            ["Impuesto_Alumbrado_Publico", "Impuesto por alumbrado público", "Decimal", "Pesos", ""],
            ["Ajuste_IAP_Otros_Meses", "Ajuste de impuesto de alumbrado público de otros meses", "Decimal", "Pesos", ""],
            ["Convivencia_Ciudadana", "Cargo por convivencia ciudadana", "Decimal", "Pesos", ""],
            ["Tasa_Especial_Convivencia", "Tasa especial de convivencia ciudadana", "Decimal", "Pesos", ""],
            ["Ajuste_Tasa_Convivencia", "Ajuste de la tasa de convivencia de otros meses", "Decimal", "Pesos", ""],
            ["Total_Servicio_Energia_Impuestos", "Total por servicio de energía e impuestos", "Decimal", "Pesos", ""],
            ["Ajuste_Decena", "Ajuste a la decena", "Decimal", "Pesos", ""],
            ["Neto_Pagar", "Valor neto a pagar", "Decimal", "Pesos", ""],
            ["Energia_Reactiva_Inductiva", "Valor de energía reactiva inductiva", "Decimal", "kVArh", ""],
            ["Energia_Reactiva_Capacitiva", "Valor de energía reactiva capacitiva", "Decimal", "kVArh", ""],
            ["Total_Energia_Reactiva", "Total de energía reactiva", "Decimal", "kVArh", ""]
        ]
        
        # Añadir metadatos para tabla Conceptos
        conceptos_metadatos = [
            ["ID_Factura", "Identificador único de la factura (clave foránea)", "UUID", "N/A", "Relaciona con tabla Facturas"],
            ["Codigo_SIC", "Código del Sistema de Información Comercial", "String", "N/A", "Identifica la cuenta en el sistema"],
            ["Concepto", "Tipo de concepto facturado", "String", "N/A", "Ej. Generación, Comercialización"],
            ["KWh_KVArh", "Cantidad de energía consumida", "Decimal", "kWh/kVArh", "Solo aplica para algunos conceptos"],
            ["Precio_KWh", "Precio unitario por kWh", "Decimal", "$/kWh", ""],
            ["Mes_Corriente", "Valor facturado del mes actual", "Decimal", "Pesos", ""],
            ["Mes_Anteriores", "Valor facturado de meses anteriores", "Decimal", "Pesos", ""],
            ["Total", "Valor total del concepto", "Decimal", "Pesos", ""]
        ]
        
        # Añadir metadatos para tabla Autorizaciones
        autorizaciones_metadatos = [
            ["ID_Factura", "Identificador único de la factura (clave foránea)", "UUID", "N/A", "Relaciona con tabla Facturas"],
            ["Codigo_SIC", "Código del Sistema de Información Comercial", "String", "N/A", "Identifica la cuenta en el sistema"],
            ["HES1", "Autorización 1", "Integer", "N/A", "Código de autorización 1"],
            ["HES2", "Autorización 2", "Integer", "N/A", "Código de autorización 2"],
            ["HES3", "Autorización 3", "Integer", "N/A", "Código de autorización 3"],
            ["HES4", "Autorización 4", "Integer", "N/A", "Código de autorización 4"],
            ["HES5", "Autorización 5", "Integer", "N/A", "Código de autorización 5"],
            ["HES6", "Autorización 6", "Integer", "N/A", "Código de autorización 6"],
            ["HES7", "Autorización 7", "Integer", "N/A", "Código de autorización 7"],
            ["HES8", "Autorización 8", "Integer", "N/A", "Código de autorización 8"],
            ["HES9", "Autorización 9", "Integer", "N/A", "Código de autorización 9"],
            ["HES10", "Autorización 10", "Integer", "N/A", "Código de autorización 10"]
        ]
        
        # Añadir metadatos para tabla Parámetros Específicos
        parametros_metadatos = [
            ["ID_Factura", "Identificador único de la factura (clave foránea)", "UUID", "N/A", "Relaciona con tabla Facturas"],
            ["Codigo_SIC", "Código del Sistema de Información Comercial", "String", "N/A", "Identifica la cuenta en el sistema"],
            ["IR", "Parámetro IR", "Decimal", "N/A", "Valor numérico del parámetro IR"],
            ["Grupo", "Número de grupo", "Decimal", "N/A", "Identificador del grupo"],
            ["DIU_INT", "Parámetro DIU INT", "Decimal", "N/A", "Valor numérico del parámetro DIU INT"],
            ["DIUM_INT", "Parámetro DIUM INT", "Decimal", "N/A", "Valor numérico del parámetro DIUM INT"],
            ["FIU_INT", "Parámetro FIU INT", "Decimal", "N/A", "Valor numérico del parámetro FIU INT"],
            ["FIUM_INT", "Parámetro FIUM INT", "Decimal", "N/A", "Valor numérico del parámetro FIUM INT"],
            ["FIUG", "Parámetro FIUG", "Decimal", "N/A", "Valor decimal del parámetro FIUG"],
            ["DIUG", "Parámetro DIUG", "Decimal", "N/A", "Valor decimal del parámetro DIUG"]
        ]
        
        # Combinar todos los metadatos
        all_metadata = filas_metadatos + conceptos_metadatos + autorizaciones_metadatos + parametros_metadatos
        
        for row_idx, metadata in enumerate(all_metadata, 2):
            for col_idx, value in enumerate(metadata, 1):
                cell = ws_metadata.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border
        
        # Ajustar ancho de columnas
        self._ajustar_ancho_columnas(ws_metadata)
    
    def _ajustar_ancho_columnas(self, worksheet):
        """
        Ajusta el ancho de las columnas de una hoja de cálculo.
        
        Args:
            worksheet: Hoja de cálculo a ajustar
        """
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width


class ExportadorExcelMultiple:
    """
    Clase para exportar múltiples facturas a un único archivo Excel.
    Permite consolidar datos de varias facturas en un solo archivo.
    """
    
    def __init__(self, ruta_salida=None):
        """
        Inicializa el exportador múltiple.
        
        Args:
            ruta_salida (str, optional): Ruta donde se guardará el archivo Excel.
                                          Si es None, se usará 'facturas_analizadas.xlsx'.
        """
        self.ruta_salida = ruta_salida or "facturas_analizadas.xlsx"
        self.workbook = Workbook()
        
        # Inicializar hojas
        self.ws_facturas = self.workbook.active
        self.ws_facturas.title = "Facturas"
        self.ws_conceptos = self.workbook.create_sheet(title="Conceptos")
        self.ws_autorizaciones = self.workbook.create_sheet(title="Autorizaciones")
        self.ws_parametros = self.workbook.create_sheet(title="Datos_OR")
        self.ws_metadata = self.workbook.create_sheet(title="Metadatos")
        
        # Definir estilos
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Contador de filas
        self.fila_facturas = 1
        self.fila_conceptos = 1
        self.fila_autorizaciones = 1
        self.fila_parametros = 1
        
        # Crear cabeceras
        self._crear_cabeceras()
    
    def _crear_cabeceras(self):
        """
        Crea las cabeceras de las hojas del Excel.
        """
        # Encabezados Facturas
        encabezados_factura = [
            "ID_Factura", "Nombre_Archivo", "Numero_Factura", "Fecha_Procesamiento", "Fecha_Vencimiento", "Periodo_Facturacion", 
            "Factor_M", "Codigo_SIC", "Subtotal_Base_Energia", "Contribucion", 
            "Contribucion_Otros_Meses", "Subtotal_Energia_Contribucion_KWh", 
            "Subtotal_Energia_Contribucion_Pesos", "Otros_Cobros", "Sobretasa", 
            "Ajustes_Cargos_Regulados", "Compensaciones", "Saldo_Cartera", 
            "Interes_Mora","Recobros", "Alumbrado_Publico", "Impuesto_Alumbrado_Publico", 
            "Ajuste_IAP_Otros_Meses", "Convivencia_Ciudadana", "Tasa_Especial_Convivencia", 
            "Ajuste_Tasa_Convivencia", "Total_Servicio_Energia_Impuestos", 
            "Ajuste_Decena", "Neto_Pagar", "Energia_Reactiva_Inductiva", 
            "Energia_Reactiva_Capacitiva", "Total_Energia_Reactiva"
        ]
        
        for col_idx, header in enumerate(encabezados_factura, 1):
            cell = self.ws_facturas.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Encabezados Conceptos
        headers_conceptos = ["ID_Factura", "Codigo_SIC", "Concepto", "KWh_KVArh", "Precio_KWh", "Mes_Corriente", "Mes_Anteriores", "Total"]
        
        for col_idx, header in enumerate(headers_conceptos, 1):
            cell = self.ws_conceptos.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Encabezados Autorizaciones
        headers_autorizaciones = [
            "ID_Factura", "Codigo_SIC", "HES1", "HES2", "HES3", "HES4", "HES5", 
            "HES6", "HES7", "HES8", "HES9", "HES10"
        ]
        
        for col_idx, header in enumerate(headers_autorizaciones, 1):
            cell = self.ws_autorizaciones.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Encabezados Parámetros Específicos
        headers_parametros = [
            "ID_Factura", "Codigo_SIC", "IR", "Grupo", "DIU_INT", "DIUM_INT", 
            "FIU_INT", "FIUM_INT", "FIUG", "DIUG"
        ]
        
        for col_idx, header in enumerate(headers_parametros, 1):
            cell = self.ws_parametros.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Actualizar contadores de filas
        self.fila_facturas = 2
        self.fila_conceptos = 2
        self.fila_autorizaciones = 2
        self.fila_parametros = 2
        
        # Crear hoja de metadatos
        self._crear_hoja_metadatos()
    
    def agregar_factura(self, datos_procesados, nombre_archivo=None):
        """
        Agrega una factura al archivo Excel.
        
        Args:
            datos_procesados (dict): Datos procesados de la factura
            nombre_archivo (str, optional): Nombre del archivo PDF original
        """
        invoice_id = str(uuid.uuid4())
        
        # Agregar datos a la hoja Facturas
        datos_generales = datos_procesados["datos_generales"]
        factura_valores = [
            invoice_id,
            nombre_archivo or "No especificado",
            datos_generales.get("numero_factura", "No encontrado"),
            datos_procesados.get("fecha_procesamiento", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            datos_generales.get("fecha_vencimiento", "No encontrado"),
            datos_generales.get("periodo_facturacion", "No encontrado"),
            datos_generales.get("factor_m", "No encontrado"),
            datos_generales.get("codigo_sic", "No encontrado"),
            datos_generales.get("subtotal_base_energia", 0),
            datos_generales.get("contribucion", 0),
            datos_generales.get("contribucion_otros_meses", 0),
            datos_generales.get("subtotal_energia_contribucion_kwh", 0),
            datos_generales.get("subtotal_energia_contribucion_pesos", 0),
            datos_generales.get("otros_cobros", 0),
            datos_generales.get("sobretasa", 0),
            datos_generales.get("ajustes_cargos_regulados", 0),
            datos_generales.get("compensaciones", 0),
            datos_generales.get("saldo_cartera", 0),
            datos_generales.get("interes_mora", 0),
            datos_generales.get("recobros", 0),
            datos_generales.get("alumbrado_publico", 0),
            datos_generales.get("impuesto_alumbrado_publico", 0),
            datos_generales.get("ajuste_iap_otros_meses", 0),
            datos_generales.get("convivencia_ciudadana", 0),
            datos_generales.get("tasa_especial_convivencia", 0),
            datos_generales.get("ajuste_tasa_convivencia", 0),
            datos_generales.get("total_servicio_energia_impuestos", 0),
            datos_generales.get("ajuste_decena", 0),
            datos_generales.get("neto_pagar", 0),
            datos_generales.get("energia_reactiva_inductiva", 0),
            datos_generales.get("energia_reactiva_capacitiva", 0),
            datos_generales.get("total_energia_reactiva", 0)
        ]
        
        for col_idx, value in enumerate(factura_valores, 1):
            cell = self.ws_facturas.cell(row=self.fila_facturas, column=col_idx, value=value)
            cell.border = self.border
        
        # Agregar datos a la hoja Conceptos
        codigo_sic = datos_generales.get("codigo_sic", "No encontrado")
        componentes = datos_procesados["componentes"]
        for componente in componentes:
            # Añadir ID de factura a cada componente
            self.ws_conceptos.cell(row=self.fila_conceptos, column=1, value=invoice_id).border = self.border
            
            # Añadir código SIC a cada componente
            self.ws_conceptos.cell(row=self.fila_conceptos, column=2, value=codigo_sic).border = self.border
            
            # Datos del componente
            self.ws_conceptos.cell(row=self.fila_conceptos, column=3, value=componente.get("concepto", "")).border = self.border
            self.ws_conceptos.cell(row=self.fila_conceptos, column=4, value=componente.get("kwh_kvarh", "N/A")).border = self.border
            self.ws_conceptos.cell(row=self.fila_conceptos, column=5, value=componente.get("precio_kwh", 0)).border = self.border
            self.ws_conceptos.cell(row=self.fila_conceptos, column=6, value=componente.get("mes_corriente", 0)).border = self.border
            self.ws_conceptos.cell(row=self.fila_conceptos, column=7, value=componente.get("mes_anteriores", 0)).border = self.border
            self.ws_conceptos.cell(row=self.fila_conceptos, column=8, value=componente.get("total", 0)).border = self.border
            
            self.fila_conceptos += 1
        
        # Agregar datos a la hoja Autorizaciones
        hes_values = [
            invoice_id,
            codigo_sic,
            datos_generales.get("hes1", ""),
            datos_generales.get("hes2", ""),
            datos_generales.get("hes3", ""),
            datos_generales.get("hes4", ""),
            datos_generales.get("hes5", ""),
            datos_generales.get("hes6", ""),
            datos_generales.get("hes7", ""),
            datos_generales.get("hes8", ""),
            datos_generales.get("hes9", ""),
            datos_generales.get("hes10", "")
        ]
        
        for col_idx, value in enumerate(hes_values, 1):
            cell = self.ws_autorizaciones.cell(row=self.fila_autorizaciones, column=col_idx, value=value)
            cell.border = self.border
        
        self.fila_autorizaciones += 1
        
        # Agregar datos a la hoja Parámetros Específicos
        parametros_especificos = datos_procesados["parametros_especificos"]
        parametros_valores = [
            invoice_id,
            codigo_sic,
            parametros_especificos.get("ir", 0),
            parametros_especificos.get("grupo", 0),
            parametros_especificos.get("diu_int", 0),
            parametros_especificos.get("dium_int", 0),
            parametros_especificos.get("fiu_int", 0),
            parametros_especificos.get("fium_int", 0),
            parametros_especificos.get("fiug", 0),
            parametros_especificos.get("diug", 0)
        ]
        
        for col_idx, value in enumerate(parametros_valores, 1):
            cell = self.ws_parametros.cell(row=self.fila_parametros, column=col_idx, value=value)
            cell.border = self.border
        
        self.fila_parametros += 1
        
        # Actualizar contador de filas de facturas
        self.fila_facturas += 1
        
        return invoice_id
    def agregar_hoja_comparacion(self, comparaciones_df):
        """
        Agrega una hoja de comparación con la base de datos.
        
        Args:
            comparaciones_df (pandas.DataFrame): DataFrame con las comparaciones
        """
        if comparaciones_df.empty:
            return
        
        # Crear hoja de comparaciones
        ws_comparaciones = self.workbook.create_sheet(title="Comparación DB")
        
        # Encabezados de la tabla de Comparaciones
        headers_comparaciones = ["ID_Factura", "Frontera", "Concepto", "Valor_Factura", 
                                "Valor_Datalake", "Diferencia%", "Estado"]
        
        for col_idx, header in enumerate(headers_comparaciones, 1):
            cell = ws_comparaciones.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Añadir datos de comparaciones
        for row_idx, row in enumerate(comparaciones_df.itertuples(), 2):
            ws_comparaciones.cell(row=row_idx, column=1, value=row.ID_Factura).border = self.border
            ws_comparaciones.cell(row=row_idx, column=2, value=row.Frontera).border = self.border
            ws_comparaciones.cell(row=row_idx, column=3, value=row.Concepto).border = self.border
            ws_comparaciones.cell(row=row_idx, column=4, value=row.Valor_Factura).border = self.border
            
            # Valor de Datalake (puede ser None)
            valor_db_cell = ws_comparaciones.cell(row=row_idx, column=5, value=row.Valor_Datalake)
            valor_db_cell.border = self.border
            
            # Diferencia (puede ser None)
            if hasattr(row, 'Diferencia') and row.Diferencia is not None:
                diferencia_cell = ws_comparaciones.cell(row=row_idx, column=6, value=row.Diferencia)
                diferencia_cell.border = self.border
                
                # Formato condicional para la diferencia
                if abs(row.Diferencia) > 1:  # Tolerancia de 1 para diferencias por redondeo
                    diferencia_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    diferencia_cell.font = Font(color="9C0006")
            else:
                ws_comparaciones.cell(row=row_idx, column=6, value=None).border = self.border
            
            # Estado
            estado_cell = ws_comparaciones.cell(row=row_idx, column=7, value=getattr(row, 'Estado', 'N/A'))
            estado_cell.border = self.border
            
            # Aplicar formato según el estado
            if getattr(row, 'Estado', '') == 'Alerta':
                estado_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                estado_cell.font = Font(color="9C0006")
            elif getattr(row, 'Estado', '') == 'No encontrado en DB':
                estado_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                estado_cell.font = Font(color="9C6500")
        
        # Ajustar ancho de columnas
        self._ajustar_ancho_columnas(ws_comparaciones)
    
    def _crear_hoja_metadatos(self):
        """
        Crea la hoja de metadatos con información descriptiva de los campos.
        """
        metadata_headers = ["Campo", "Descripción", "Tipo de Dato", "Unidad", "Observaciones"]
        
        for col_idx, header in enumerate(metadata_headers, 1):
            cell = self.ws_metadata.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Añadir descripciones de metadatos para tabla Facturas
        filas_metadatos = [
            ["ID_Factura", "Identificador único de la factura", "UUID", "N/A", "Clave primaria"],
            ["Nombre_Archivo", "Nombre del archivo PDF de la factura", "String", "N/A", "Nombre de archivo original"],
            ["Numero_Factura", "Número de la factura electrónica", "String", "N/A", "Ej. E1888"],
            ["Fecha_Procesamiento", "Fecha y hora de procesamiento del archivo", "Datetime", "N/A", "Generado automáticamente"],
            ["Fecha_Vencimiento", "Fecha límite de pago", "Date", "N/A", ""],
            ["Periodo_Facturacion", "Período al que corresponde la factura", "String", "N/A", ""],
            ["Factor_M", "Factor de multiplicación para mediciones", "Integer", "N/A", ""],
            ["Codigo_SIC", "Código del Sistema de Información Comercial", "String", "N/A", ""],
            ["Subtotal_Base_Energia", "Subtotal base de energía facturada", "Decimal", "Pesos", ""],
            ["Contribucion", "Valor de la contribución", "Decimal", "Pesos", ""],
            ["Contribucion_Otros_Meses", "Contribución correspondiente a meses anteriores", "Decimal", "Pesos", ""],
            ["Subtotal_Energia_Contribucion_KWh", "Precio por kWh de energía y contribución", "Decimal", "$/kWh", ""],
            ["Subtotal_Energia_Contribucion_Pesos", "Subtotal energía más contribución", "Decimal", "Pesos", ""],
            ["Otros_Cobros", "Valor de otros cobros adicionales", "Decimal", "Pesos", ""],
            ["Sobretasa", "Valor de la sobretasa aplicada", "Decimal", "Pesos", ""],
            ["Ajustes_Cargos_Regulados", "Ajustes por cargos regulados", "Decimal", "Pesos", ""],
            ["Compensaciones", "Valor de compensaciones aplicadas", "Decimal", "Pesos", ""],
            ["Saldo_Cartera", "Saldo pendiente en cartera", "Decimal", "Pesos", ""],
            ["Interes_Mora", "Interés por mora en pagos anteriores", "Decimal", "Pesos", ""],
            ["Recobros", "Trabajos otc", "Decimal", "Pesos", ""],
            ["Alumbrado_Publico", "Cargo por alumbrado público", "Decimal", "Pesos", ""],
            ["Impuesto_Alumbrado_Publico", "Impuesto por alumbrado público", "Decimal", "Pesos", ""],
            ["Ajuste_IAP_Otros_Meses", "Ajuste de impuesto de alumbrado público de otros meses", "Decimal", "Pesos", ""],
            ["Convivencia_Ciudadana", "Cargo por convivencia ciudadana", "Decimal", "Pesos", ""],
            ["Tasa_Especial_Convivencia", "Tasa especial de convivencia ciudadana", "Decimal", "Pesos", ""],
            ["Ajuste_Tasa_Convivencia", "Ajuste de la tasa de convivencia de otros meses", "Decimal", "Pesos", ""],
            ["Total_Servicio_Energia_Impuestos", "Total por servicio de energía e impuestos", "Decimal", "Pesos", ""],
            ["Ajuste_Decena", "Ajuste a la decena", "Decimal", "Pesos", ""],
            ["Neto_Pagar", "Valor neto a pagar", "Decimal", "Pesos", ""],
            ["Energia_Reactiva_Inductiva", "Valor de energía reactiva inductiva", "Decimal", "kVArh", ""],
            ["Energia_Reactiva_Capacitiva", "Valor de energía reactiva capacitiva", "Decimal", "kVArh", ""],
            ["Total_Energia_Reactiva", "Total de energía reactiva", "Decimal", "kVArh", ""]
        ]
        
        # Añadir metadatos para tabla Conceptos
        conceptos_metadatos = [
            ["ID_Factura", "Identificador único de la factura (clave foránea)", "UUID", "N/A", "Relaciona con tabla Facturas"],
            ["Codigo_SIC", "Código del Sistema de Información Comercial", "String", "N/A", "Identifica la cuenta en el sistema"],
            ["Concepto", "Tipo de concepto facturado", "String", "N/A", "Ej. Generación, Comercialización"],
            ["KWh_KVArh", "Cantidad de energía consumida", "Decimal", "kWh/kVArh", "Solo aplica para algunos conceptos"],
            ["Precio_KWh", "Precio unitario por kWh", "Decimal", "$/kWh", ""],
            ["Mes_Corriente", "Valor facturado del mes actual", "Decimal", "Pesos", ""],
            ["Mes_Anteriores", "Valor facturado de meses anteriores", "Decimal", "Pesos", ""],
            ["Total", "Valor total del concepto", "Decimal", "Pesos", ""]
        ]
        
        # Añadir metadatos para tabla Autorizaciones
        autorizaciones_metadatos = [
            ["ID_Factura", "Identificador único de la factura (clave foránea)", "UUID", "N/A", "Relaciona con tabla Facturas"],
            ["Codigo_SIC", "Código del Sistema de Información Comercial", "String", "N/A", "Identifica la cuenta en el sistema"],
            ["HES1", "Autorización 1", "Integer", "N/A", "Código de autorización 1"],
            ["HES2", "Autorización 2", "Integer", "N/A", "Código de autorización 2"],
            ["HES3", "Autorización 3", "Integer", "N/A", "Código de autorización 3"],
            ["HES4", "Autorización 4", "Integer", "N/A", "Código de autorización 4"],
            ["HES5", "Autorización 5", "Integer", "N/A", "Código de autorización 5"],
            ["HES6", "Autorización 6", "Integer", "N/A", "Código de autorización 6"],
            ["HES7", "Autorización 7", "Integer", "N/A", "Código de autorización 7"],
            ["HES8", "Autorización 8", "Integer", "N/A", "Código de autorización 8"],
            ["HES9", "Autorización 9", "Integer", "N/A", "Código de autorización 9"],
            ["HES10", "Autorización 10", "Integer", "N/A", "Código de autorización 10"]
        ]
        
        # Añadir metadatos para tabla Parámetros Específicos
        parametros_metadatos = [
            ["ID_Factura", "Identificador único de la factura (clave foránea)", "UUID", "N/A", "Relaciona con tabla Facturas"],
            ["Codigo_SIC", "Código del Sistema de Información Comercial", "String", "N/A", "Identifica la cuenta en el sistema"],
            ["IR", "Parámetro IR", "Decimal", "N/A", "Valor numérico del parámetro IR"],
            ["Grupo", "Número de grupo", "Decimal", "N/A", "Identificador del grupo"],
            ["DIU_INT", "Parámetro DIU INT", "Decimal", "N/A", "Valor numérico del parámetro DIU INT"],
            ["DIUM_INT", "Parámetro DIUM INT", "Decimal", "N/A", "Valor numérico del parámetro DIUM INT"],
            ["FIU_INT", "Parámetro FIU INT", "Decimal", "N/A", "Valor numérico del parámetro FIU INT"],
            ["FIUM_INT", "Parámetro FIUM INT", "Decimal", "N/A", "Valor numérico del parámetro FIUM INT"],
            ["FIUG", "Parámetro FIUG", "Decimal", "N/A", "Valor decimal del parámetro FIUG"],
            ["DIUG", "Parámetro DIUG", "Decimal", "N/A", "Valor decimal del parámetro DIUG"]
        ]
        
        
        # Combinar todos los metadatos
        all_metadata = filas_metadatos + conceptos_metadatos + autorizaciones_metadatos + parametros_metadatos
        
        for row_idx, metadata in enumerate(all_metadata, 2):
            for col_idx, value in enumerate(metadata, 1):
                cell = self.ws_metadata.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border
    
    def finalizar(self):
        """
        Finaliza el proceso de exportación y guarda el archivo Excel.
        
        Returns:
            str: Ruta del archivo Excel creado
        """
        # Ajustar ancho de columnas
        for worksheet in [self.ws_facturas, self.ws_conceptos, self.ws_autorizaciones, self.ws_parametros, self.ws_metadata]:
            self._ajustar_ancho_columnas(worksheet)
        
        # Guardar archivo
        self.workbook.save(self.ruta_salida)
        return self.ruta_salida
    
    def _ajustar_ancho_columnas(self, worksheet):
        """
        Ajusta el ancho de las columnas de una hoja de cálculo.
        
        Args:
            worksheet: Hoja de cálculo a ajustar
        """
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width


def procesar_multiples_facturas(directorio_entrada, ruta_salida=None, directorio_csv=None, fecha_inicio=None, fecha_fin=None):
    """
    Procesa múltiples facturas en un directorio y las exporta a un solo archivo Excel.
    
    Args:
        directorio_entrada (str): Directorio con los archivos PDF de facturas
        ruta_salida (str, optional): Ruta donde se guardará el archivo Excel.
                                     Si es None, se usará 'facturas_analizadas.xlsx'.
        directorio_csv (str, optional): Directorio donde se guardarán los archivos CSV.
                                        Si es None, se usará 'csv' dentro del directorio de salida.
        fecha_inicio (str, optional): Fecha de inicio para filtrar las facturas en BD (YYYY-MM-DD)
        fecha_fin (str, optional): Fecha de fin para filtrar las facturas en BD (YYYY-MM-DD)
    
    Returns:
        str: Ruta del archivo Excel creado
    """
    import extractores
    import procesamiento
    from db_connector import DBConnector
    import pandas as pd
    
    # Definir ruta de salida
    if ruta_salida is None:
        ruta_salida = os.path.join(os.path.dirname(directorio_entrada), "resultados")
    
    # Crear directorio de salida si no existe
    if not os.path.exists(ruta_salida):
        os.makedirs(ruta_salida)
    
    # Definir directorio para archivos CSV
    if directorio_csv is None:
        directorio_csv = os.path.join(ruta_salida, "csv")
    
    # Crear directorio para archivos CSV si no existe
    if not os.path.exists(directorio_csv):
        os.makedirs(directorio_csv)
    
    # Ruta del archivo Excel consolidado
    ruta_excel = os.path.join(ruta_salida, "facturas_analizadas.xlsx")
    
    # Obtener lista de archivos PDF
    archivos_pdf = [f for f in os.listdir(directorio_entrada) if f.lower().endswith('.pdf')]
    
    if not archivos_pdf:
        print(f"No se encontraron archivos PDF en el directorio {directorio_entrada}")
        return None
    
    # Crear exportador de Excel consolidado
    exportador = ExportadorExcelMultiple(ruta_excel)
    
    # Lista para guardar datos procesados para la comparación
    facturas_procesadas = []
    
    # Procesar cada factura
    for archivo in archivos_pdf:
        ruta_pdf = os.path.join(directorio_entrada, archivo)
        
        # Validar archivo
        if not procesamiento.validar_ruta_archivo(ruta_pdf):
            print(f"Omitiendo archivo: {archivo}")
            continue
        
        # Extraer nombre base
        nombre_base = os.path.splitext(archivo)[0]
        
        # Convertir PDF a CSV (guardándolo en el directorio de CSV)
        ruta_csv = os.path.join(directorio_csv, f"{nombre_base}.csv")
        extractores.convertir_pdf_a_csv(ruta_pdf, ruta_csv)
        
        # Extraer datos del CSV
        datos_generales, datos_componentes = extractores.extraer_todos_datos_factura(ruta_csv)
        
        # Procesar datos
        processor = procesamiento.FacturaProcessor(datos_generales, datos_componentes)
        datos_procesados = processor.obtener_datos_procesados()
        
        # Agregar a Excel consolidado incluyendo el nombre del archivo
        factura_id = exportador.agregar_factura(datos_procesados, archivo)
        
        # Guardar el ID en los datos procesados para la comparación
        datos_procesados['id'] = factura_id
        facturas_procesadas.append(datos_procesados)
    
    if facturas_procesadas:
        # Imprimir información de diagnóstico
        for factura in facturas_procesadas:
            print(f"Procesada factura con código SIC: {factura['datos_generales'].get('codigo_sic')}")
            print(f"Período de facturación: {factura['datos_generales'].get('periodo_facturacion')}")
            print("----")
    
        # Realizar comparación con la base de datos
        try:
            # Usar el conector de base de datos
            db_connector = DBConnector()
            
            # Comparar con la base de datos usando fechas proporcionadas o extraídas de las facturas
            comparaciones_df = db_connector.compare_with_facturas(
                facturas_procesadas,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin
            )
            
            # Agregar hoja de comparación si hay datos
            if not comparaciones_df.empty:
                exportador.agregar_hoja_comparacion(comparaciones_df)
                print(f"Se agregó hoja de comparación con {len(comparaciones_df)} registros")
            else:
                print("No se encontraron datos para comparación con la base de datos")
        
        except Exception as e:
            print(f"Error al comparar con la base de datos: {e}")
            import traceback
            traceback.print_exc()
    
    # Finalizar y guardar Excel
    ruta_final = exportador.finalizar()
    print(f"Archivo Excel consolidado creado: {ruta_final}")
    
    return ruta_final